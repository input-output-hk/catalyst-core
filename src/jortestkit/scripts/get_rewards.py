import contextlib
import os
import time

import openpyxl
import requests
from requests.exceptions import HTTPError
import pandas as pd
from openpyxl import load_workbook

api_url_base = "http://127.0.0.1:9001/api"
reward_parameters_constant = 3835616440000
reward_drawing_limit_max = 4109589/10000000000
treasury_parameters_ratio = 1/10
slot_duration = 1
slots_per_epoch = 60
epoch_time_secs = slot_duration * slots_per_epoch
filename = 'rewards_history.xlsx'
log_file = 'rewards_history.txt'

api_url = f"{api_url_base}/v0"

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.max_colwidth', -1)


def wait_for_new_epoch():
    current_epoch, _ = get_stake_pools_current_epoch()
    new_epoch = current_epoch + 1
    print(f"=== waiting for epoch {current_epoch} to end ...")
    counter = 0
    while current_epoch != new_epoch:
        time.sleep(slot_duration)
        current_epoch, _ = get_stake_pools_current_epoch()
        if counter > epoch_time_secs + 5:
            print(f"!!! ERROR: No new epoch in {epoch_time_secs + 5} seconds")
            exit(2)


def endpoint(url):
    try:
        r = requests.get(url)
        r.raise_for_status()
    except HTTPError as http_err:
        print("\nWeb API unavailable.\nError Details:\n")
        print(f"HTTP error occurred: {http_err}")
        exit(1)
    except Exception as err:
        print("\nWeb API unavailable.\nError Details:\n")
        print(f"Other error occurred: {err}")
        exit(1)
    else:
        return(r)


def get_api(path):
    r = endpoint(f'{api_url}/{path}')
    return r.text


def get_tip():
    return get_api("tip")


def get_block(block_id):
    r = endpoint(f'{api_url}/block/{block_id}')
    hex_block = r.content.hex()
    return hex_block


def get_leader_logs():
    r = endpoint(f'{api_url}/leaders/logs')
    return r.json()


def get_stake():
    r = endpoint(f'{api_url}/stake')
    return r.json()


def get_stake_pool(stake_pool_id):
    r = endpoint(f'{api_url}/stake_pool/{stake_pool_id}')
    return r.json()


def parse_block(block):
    return {
        "epoch": int(block[16:24], 16),
        "slot": int(block[24:32], 16),
        "parent": block[104:168],
        "pool": block[168:232],
    }


def get_official_history(no_of_epochs):
    block_hash = get_tip()
    block = parse_block(get_block(block_hash))
    currentEpoch = block['epoch']
    row_list = []

    while block['epoch'] > currentEpoch - no_of_epochs:
        row_dict = {'date': str(block['epoch']) + '.' + str(block['slot']),
                    'block_hash': block_hash,
                    'parent_hash': block['parent'],
                    'pool_id': block['pool']}
        row_list.append(row_dict)

        block_hash = block['parent']
        block = parse_block(get_block(block['parent']))

    df_official = pd.DataFrame(row_list, columns=['date', 'block_hash', 'parent_hash', 'pool_id'])

    return df_official


def get_no_of_blocks_for_epoch(epoch_no):
    block_hash = get_tip()
    block = parse_block(get_block(block_hash))
    currentEpoch = block['epoch']

    official_history_df = get_official_history(currentEpoch - epoch_no + 1)
    return official_history_df['date'].str.contains(str(epoch_no) + ".").sum()


def get_created_block_dates_per_stake_pool_per_epoch(epoch_no):
    block_hash = get_tip()
    block = parse_block(get_block(block_hash))
    currentEpoch = block['epoch']

    official_history_df = get_official_history(currentEpoch - epoch_no + 1)
    df = official_history_df.loc[official_history_df['date'].str.contains(str(epoch_no) + ".")]
    return df.groupby('pool_id')['date'].apply(list)


def get_node_history(node_port, no_of_epochs):
    global row_dict
    row_list = []
    leader_logs = get_leader_logs()
    for log in leader_logs:
        if 'Block' in log['status']:
            row_dict = {'created_at_time': log['created_at_time'],
                        'scheduled_at_time': log['scheduled_at_time'],
                        'scheduled_at_date': log['scheduled_at_date'],
                        'wake_at_time': log['wake_at_time'],
                        'finished_at_time': log['finished_at_time'],
                        'status': 'InABlock',
                        'block': log['status']['Block']['block'],
                        'chain_length': log['status']['Block']['chain_length'],
                        'enclave_leader_id': log['enclave_leader_id']}
        else:
            row_dict = {'created_at_time': log['created_at_time'],
                        'scheduled_at_time': log['scheduled_at_time'],
                        'scheduled_at_date': log['scheduled_at_date'],
                        'wake_at_time': log['wake_at_time'],
                        'finished_at_time': log['finished_at_time'],
                        'status': log['status'],
                        'block': None,
                        'chain_length': None,
                        'enclave_leader_id': log['enclave_leader_id']}
        row_list.append(row_dict)

    df_node = pd.DataFrame(row_list, columns=['created_at_time', 'scheduled_at_time', 'scheduled_at_date',
                                              'wake_at_time', 'finished_at_time', 'status', 'block', 'chain_length',
                                              'enclave_leader_id'])
    return df_node


def get_total_stake_pools_rewards_current_epoch():
    stake_per_epoch = get_stake()

    current_epoch = stake_per_epoch['epoch']
    current_dangling = stake_per_epoch['stake']['dangling']
    stake_pools = stake_per_epoch['stake']['pools']
    unassigned = stake_per_epoch['stake']['unassigned']

    print(f"=== current_epoch   : {current_epoch}")
    sum_stake = 0
    sum_value_for_stakers = 0
    sum_value_taxed = 0
    for pool in stake_pools:
        sum_stake += pool[1]

        stake_pool_details = get_stake_pool(pool[0])
        sum_value_for_stakers += stake_pool_details['rewards']['value_for_stakers']
        sum_value_taxed += stake_pool_details['rewards']['value_taxed']

    # print(f"sum_stake                     : {sum_stake}")
    # print(f"total_rewards_for_stakers     : {sum_value_for_stakers}")
    # print(f"total_rewards_for_stake_pools : {sum_value_taxed}")
    print(f"total_rewards_per_epoch    : {sum_value_for_stakers + sum_value_taxed}")
    print(f"reward_parameters_constant : {reward_parameters_constant}")
    print(f"dif_expected_real          : {sum_value_for_stakers + sum_value_taxed - reward_parameters_constant}")
    print(f"===========================================================================")


def get_stake_pools_current_epoch():
    stake_per_epoch = get_stake()

    current_epoch = stake_per_epoch['epoch']
    current_dangling = stake_per_epoch['stake']['dangling']
    stake_pools = stake_per_epoch['stake']['pools']
    unassigned = stake_per_epoch['stake']['unassigned']

    return current_epoch, stake_pools


def fill_rewards_history_file(measurement_no_of_epochs):
    # delete the excel file if it exists
    global row_no
    with contextlib.suppress(FileNotFoundError):
        os.remove(filename)
        os.remove(log_file)

    # create the excel file
    workbook = openpyxl.Workbook()
    workbook.save(filename=filename)

    # create the logging file
    with open(log_file, 'a') as f:
        f.close()

    #  open the excel file
    workbook = load_workbook(filename=filename)

    #  change the name of the initial worksheet
    active_worksheet = workbook['Sheet']
    active_worksheet.title = 'rewards_per_epoch'

    # Append column names to first worksheet
    active_worksheet.append(["epoch_no", "treasury_rewards_per_epoch", "stake_pools_total_rewards_per_epoch",
                             "reward_parameters_constant", "diff_expected_real", "no_of_blocks_per_epoch",
                             "total_active_stake", "reward_drawing_limit_max", "no_of_txs_per_epoch", "rewards_per_block"])

    # create the second worksheet
    workbook.create_sheet('rewards_per_stake_pool')

    # open the second worksheet
    active_worksheet = workbook['rewards_per_stake_pool']

    # Append column names to second worksheet
    active_worksheet.append(["epoch_no", "stake_pool_id", "no_of_created_blocks", "value_for_stakers", "value_taxed",
                             "total_rewards_per_stake_pool", "rewards_per_block", "stake", "tax_fixed", "tax_max",
                             "tax_ratio_denominator", "tax_ratio_numerator"])

    for counter in range(measurement_no_of_epochs):
        # get the stake pool details - from `jcli rest v0 stake get`
        current_epoch, stake_pools = get_stake_pools_current_epoch()

        # open the `rewards_per_stake_pool` worksheet
        active_worksheet = workbook["rewards_per_stake_pool"]
        header_list = [cell.value for cell in active_worksheet[1]]

        with open(log_file, 'a') as f:
            f.writelines(f"==== current_epoch: {current_epoch}\n")
            f.close()

        for stake_pool in stake_pools:
            row_no = active_worksheet.max_row + 1
            stake_pool_details = get_stake_pool(stake_pool[0])

            with open(log_file, 'a') as f:
                f.writelines(f"stake_pool:{stake_pool}")
                f.writelines(f"stake_pool_details:{stake_pool_details}")
                f.writelines(f"\n")
                f.close()

            #  fill the epoch_no column
            active_worksheet.cell(column=header_list.index("epoch_no") + 1, row=row_no, value=current_epoch)

            #  fill the stake_pool_id column
            active_worksheet.cell(column=header_list.index("stake_pool_id") + 1, row=row_no, value=stake_pool[0])

            #  fill the stake column
            active_worksheet.cell(column=header_list.index("stake") + 1, row=row_no, value=stake_pool[1])

            #  fill the tax_fixed column
            active_worksheet.cell(column=header_list.index("tax_fixed") + 1, row=row_no, value=stake_pool_details['tax']['fixed'])

            #  fill the tax_max column
            if "max" in stake_pool_details['tax']:
                tax_max = stake_pool_details['tax']['max']
            else:
                tax_max = 0
            active_worksheet.cell(column=header_list.index("tax_max") + 1, row=row_no, value=tax_max)

            #  fill the tax_ratio_denominator column
            active_worksheet.cell(column=header_list.index("tax_ratio_denominator") + 1, row=row_no, value=stake_pool_details['tax']['ratio']['denominator'])

            #  fill the tax_ratio_numerator column
            active_worksheet.cell(column=header_list.index("tax_ratio_numerator") + 1, row=row_no, value=stake_pool_details['tax']['ratio']['numerator'])

            if current_epoch == stake_pool_details['rewards']['epoch']:
                value_for_stakers = stake_pool_details['rewards']['value_for_stakers']
                value_taxed = stake_pool_details['rewards']['value_taxed']
            else:
                value_for_stakers = 0
                value_taxed = 0

            #  fill the value_for_stakers column
            active_worksheet.cell(column=header_list.index("value_for_stakers") + 1, row=row_no, value=value_for_stakers)

            #  fill the value_taxed column
            active_worksheet.cell(column=header_list.index("value_taxed") + 1, row=row_no, value=value_taxed)

            #  fill the total_rewards_per_stake_pool column
            col1_value = active_worksheet.cell(row=row_no, column=header_list.index("value_for_stakers") + 1).value
            col2_value = active_worksheet.cell(row=row_no, column=header_list.index("value_taxed") + 1).value
            total_rewards_per_stake_pool = col1_value + col2_value
            active_worksheet.cell(column=header_list.index("total_rewards_per_stake_pool") + 1, row=row_no, value=total_rewards_per_stake_pool)

        #  get block_dates_per_epoch created by each stake pool
        block_dates_per_epoch_df = get_created_block_dates_per_stake_pool_per_epoch(current_epoch - 1)

        with open(log_file, 'a') as f:
            f.writelines(f"==== block_dates_per_epoch_df for epoch {current_epoch - 1}\n {block_dates_per_epoch_df}")
            f.writelines(f"\n")
            f.close()

        index = 0
        # iterate thru all the stake pools that created blocks in the specified epoch
        for stake_pool in block_dates_per_epoch_df.index:
            for i in range(1, active_worksheet.max_row + 1):
                # search for the rows with the specified epoch
                # if active_worksheet.cell(row=i, column=header_list.index('epoch_no') + 1).value == current_epoch:
                #     if active_worksheet.cell(row=i, column=header_list.index('stake_pool_id') + 1).value == stake_pool:
                #         # fill the no_of_created_blocks = 0 for each stake pool for the current epoch
                #         no_of_created_blocks = 0
                #         # fill the no_of_created_blocks for each stake pool (for the specific epoch - the previous one)
                #         active_worksheet.cell(column=header_list.index('no_of_created_blocks') + 1, row=i, value=no_of_created_blocks)
                if active_worksheet.cell(row=i, column=header_list.index('epoch_no') + 1).value == current_epoch - 1:
                    # search for the row containing the specified epoch and specified stake pool
                    if active_worksheet.cell(row=i, column=header_list.index('stake_pool_id') + 1).value == stake_pool:
                        no_of_created_blocks = len(block_dates_per_epoch_df[index])
                        # fill the no_of_created_blocks for each stake pool (for the specific epoch - the previous one)
                        active_worksheet.cell(column=header_list.index('no_of_created_blocks') + 1, row=i, value=no_of_created_blocks)
            index += 1

        # read the total_rewards per epoch in dictionary
        total_rewards_per_epoch_dict = {}
        for i in range(2, active_worksheet.max_row + 1):
            epoch = active_worksheet.cell(row=i, column=header_list.index('epoch_no') + 1).value
            total_rewards_per_sp = active_worksheet.cell(row=i, column=header_list.index('total_rewards_per_stake_pool') + 1).value
            if epoch not in total_rewards_per_epoch_dict:
                total_rewards_per_epoch_dict[epoch] = []
            total_rewards_per_epoch_dict[epoch].append(total_rewards_per_sp)

        # read the total_stake per epoch in dictionary
        total_stake_per_epoch_dict = {}
        for i in range(2, active_worksheet.max_row + 1):
            epoch = active_worksheet.cell(row=i, column=header_list.index('epoch_no') + 1).value
            total_stake_per_sp = active_worksheet.cell(row=i, column=header_list.index('stake') + 1).value
            if epoch not in total_stake_per_epoch_dict:
                total_stake_per_epoch_dict[epoch] = []
            total_stake_per_epoch_dict[epoch].append(total_stake_per_sp)
            
        # open the first worksheet
        active_worksheet = workbook['rewards_per_epoch']
        header_list = [cell.value for cell in active_worksheet[1]]
        row_no = active_worksheet.max_row + 1

        #  fill the epoch_no column
        active_worksheet.cell(column=header_list.index("epoch_no") + 1, row=row_no, value=current_epoch)

        # #  fill the no_of_blocks_per_epoch column
        no_of_blocks_for_epoch_val = get_no_of_blocks_for_epoch(current_epoch - 1)

        if row_no > 2:
            active_worksheet.cell(column=header_list.index("no_of_blocks_per_epoch") + 1, row=row_no-1, value=no_of_blocks_for_epoch_val)
        active_worksheet.cell(column=header_list.index("no_of_blocks_per_epoch") + 1, row=row_no, value=0)

        #  fill the reward_parameters_constant column
        active_worksheet.cell(column=header_list.index("reward_parameters_constant") + 1, row=row_no, value=reward_parameters_constant)

        #  fill the stake_pools_total_rewards_per_epoch column
        stake_pools_total_rewards_per_epoch = sum(total_rewards_per_epoch_dict.get(current_epoch))
        active_worksheet.cell(column=header_list.index("stake_pools_total_rewards_per_epoch") + 1, row=row_no, value=stake_pools_total_rewards_per_epoch)

        #  fill the treasury_rewards_per_epoch column
        treasury_rewards_per_epoch = stake_pools_total_rewards_per_epoch * 100 / 90 * treasury_parameters_ratio
        active_worksheet.cell(column=header_list.index("treasury_rewards_per_epoch") + 1, row=row_no, value=treasury_rewards_per_epoch)

        #  fill the diff_expected_real column
        diff_expected_real = reward_parameters_constant - stake_pools_total_rewards_per_epoch - treasury_rewards_per_epoch
        active_worksheet.cell(column=header_list.index("diff_expected_real") + 1, row=row_no, value=diff_expected_real)

        #  fill the total_active_stake column
        total_active_stake = sum(total_stake_per_epoch_dict.get(current_epoch))
        active_worksheet.cell(column=header_list.index("total_active_stake") + 1, row=row_no, value=total_active_stake)

        #  fill the treasury_rewards_per_epoch column
        reward_drawing_limit_max_val = total_active_stake * reward_drawing_limit_max
        active_worksheet.cell(column=header_list.index("reward_drawing_limit_max") + 1, row=row_no, value=reward_drawing_limit_max_val)

        with open(log_file, 'a') as f:
            f.writelines(f"==== get_no_of_blocks_for_epoch for epoch {current_epoch - 1} - {no_of_blocks_for_epoch_val}")
            f.writelines(f"\n")
            f.writelines(f"=================================================================================\n\n")
            f.close()

        workbook.save(filename)

        # wait for the new epoch
        wait_for_new_epoch()

        counter += 1


if __name__ == "__main__":
    # get_total_stake_pools_rewards_current_epoch()
    fill_rewards_history_file(12)
    # print(get_created_block_dates_per_stake_pool_per_epoch(31))
    # print(get_no_of_blocks_for_epoch(31))

"""Utilities to manage data."""
from dataclasses import dataclass
import json
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries
from datetime import datetime
import aiohttp
from loguru import logger
import re
from typing import List, Dict, TypeVar, Type, Any, Mapping, Optional


T = TypeVar("T")
DictOrList = TypeVar("DictOrList", Dict[str, Any], List[Any])


def load_csv_and_serialize(path: str, model: Type[T], extra: Dict) -> List[T]:
    """Given a path of csv and a model it returns a list of models, merging it with extra attributes."""
    with open(path) as f:
        data = csv.DictReader(f)
        serialized = [model(**el, **extra) for el in data]
        return serialized

def deserialize_and_save_csv(path: str, elements: List[Any], include: dict, transform: Optional[str] = None):
    """Deserialize a list of models and save a csv."""
    results = []
    for el in elements:
        res = el.dict(include=include)
        _res = {}
        if transform == "pa_allocation":
            _res["Email Address"] = res["email"]
            _res["Display Name"] = res["name"]
            _res[
                "Would you like to participate as a Reviewer in the Community Review stage?"
            ] = "Yes, I want to be a Community Reviewer and I also understand the role."
            _res["Level (DO NOT EDIT)"] = f"LVL{res['level']}"
            _res["Allocations (DO NOT EDIT)"] = "<br />".join(
                [f"- <a href=\"{p['proposal']['url']}\" target=\"_blank\">{p['proposal']['title']}</a>" for p in res["allocations"]]
            )
        elif transform == "single_allocation":
            _res = {
                **{f"pa_{key}": val for key, val in res["pa"].items()},
                **{f"proposal_{key}": val for key, val in res["proposal"].items()},
            }
        elif transform == "postprocessed_reviews":
            _res = {
                **res,
                **{f"proposal_{key}": val for key, val in res["proposal"].items()},
            }
            _res.pop("proposal")
        elif transform == "profanity":
            _res = {
                **res,
                **{f"review_{key}": val for key, val in res["review"].items()},
            }
            _res.pop("review")
        elif transform == "ai":
            _res = {
                **res,
                **{"review": res["review"]['id']},
            }
        elif transform == "similarity":
            _res = {
                **res,
                **{
                    "left": res["left"]['id'],
                    "right": res["right"]['id']
                }
            }
        else:
            _res = res

        results.append(_res)

    save_csv(path, results)
    

def save_csv(path: str, elements: List[Any]):
    keys = elements[0].keys()
    if path.endswith("csv"):
        with open(path, "w", newline="") as f:
            dict_writer = csv.DictWriter(f, keys)
            dict_writer.writeheader()
            dict_writer.writerows(elements)
    elif path.endswith("xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append(list(keys))
        for el in elements[0:]:
            ws.append(list(el.values()))
        wb.save(path)


def snake_case_keys(x: DictOrList):
    """Recursively transforms all dict keys to snake_case."""
    if isinstance(x, dict):
        keys = list(x.keys())
        for k in keys:
            v = x.pop(k)
            snake_case_keys(v)
            x[snake_case(k)] = v
    elif isinstance(x, list):
        for i in range(len(x)):
            snake_case_keys(x[i])


def snake_case(s: str) -> str:
    """Transform a string to snake_case."""
    return re.sub(r"([a-z])([A-Z])", r"\1_\2", s).lower()


def slugify(s):
    """Create a slug for filenames from a string."""
    s = s.lower().strip()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[\s_-]+", "-", s)
    s = re.sub(r"^-+|-+$", "", s)
    return s


def unmerge_xlsx(path: str) -> Workbook:
    """Unmerge cells from an xlsx file."""
    wb = load_workbook(path)
    for st_name in wb.sheetnames:
        # Unmerge cells
        st = wb[st_name]
        mcr_coord_list = [mcr.coord for mcr in st.merged_cells.ranges]

        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            top_left_cell_value = st.cell(row=min_row, column=min_col).value
            st.unmerge_cells(mcr)
            for row in st.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value
    wb.save(path.replace(".xlsx", "_unmerged.xlsx"))
    return wb


def get_rows_from_xlsx(wb, sheetname) -> List[dict]:
    """Extract rows from an xlsx file."""
    st = wb[sheetname]
    rows = []
    keys = []
    for idx, row in enumerate(st.iter_rows(values_only=True)):
        if idx == 0:
            keys = row
        else:
            rows.append(dict(zip(keys, row)))
    return rows


@dataclass
class RequestProgressInfo:
    """Information about a request's progress."""

    method: str
    url: str
    bytes_received: int
    last_update: datetime


class RequestProgressObserver:
    """Observer used for displaying IdeaScale client's requests progresses."""

    def __init__(self):
        """Initialize a new instance of RequestProgressObserver."""
        self.inflight_requests: Dict[int, RequestProgressInfo] = {}

    def request_start(self, req_id: int, method: str, url: str):
        """Register the start of a request."""
        logger.info("Request started", req_id=req_id, method=method, url=url)
        self.inflight_requests[req_id] = RequestProgressInfo(method, url, 0, datetime.now())

    def request_progress(self, req_id: int, bytes_received: int):
        """Register the progress of a request."""
        info = self.inflight_requests[req_id]
        info.bytes_received += bytes_received

        now = datetime.now()
        if (now - info.last_update).total_seconds() > 1:
            logger.debug(
                "Request progress",
                req_id=req_id,
                method=info.method,
                url=info.url,
                bytes_received=info.bytes_received,
            )
        info.last_update = now

    def request_end(self, req_id):
        """Register the end of a request."""
        info = self.inflight_requests[req_id]
        logger.info(
            "Request finished",
            req_id=req_id,
            method=info.method,
            url=info.url,
            bytes_received=info.bytes_received,
        )
        del self.inflight_requests[req_id]


class JsonHttpClient:
    """HTTP Client for JSON APIs."""

    def __init__(self, api_url: str):
        """Initialize a new instance of JsonHttpClient."""
        self.api_url = api_url
        self.request_progress_observer = RequestProgressObserver()
        self.request_counter = 0

    async def post(self, path: str, data: Mapping[str, str] = {}, headers: Mapping[str, str] = {}):
        """Execute a POST request against a service."""
        url = f"{self.api_url}{path}"

        self.request_counter += 1
        req_id = self.request_counter

        async with aiohttp.ClientSession() as session:
            self.request_progress_observer.request_start(req_id, "POST", url)
            async with session.post(url, json=data, headers=headers) as r:
                content = b""

                async for c, _ in r.content.iter_chunks():
                    content += c
                    self.request_progress_observer.request_progress(req_id, len(c))

                self.request_progress_observer.request_end(req_id)

                if r.status == 200:
                    parsed_json = json.loads(content)
                    snake_case_keys(parsed_json)
                    return parsed_json
                else:
                    raise GetFailed(r.status, r.reason, content)

    async def get(self, path: str, headers: Mapping[str, str] = {}):
        """Execute a GET request against a service."""
        url = f"{self.api_url}{path}"

        self.request_counter += 1
        req_id = self.request_counter

        async with aiohttp.ClientSession() as session:
            self.request_progress_observer.request_start(req_id, "POST", url)
            async with session.get(url, headers=headers) as r:
                content = b""

                async for c, _ in r.content.iter_chunks():
                    content += c
                    self.request_progress_observer.request_progress(req_id, len(c))

                self.request_progress_observer.request_end(req_id)

                if r.status == 200:
                    parsed_json = json.loads(content)
                    snake_case_keys(parsed_json)
                    return parsed_json
                else:
                    raise GetFailed(r.status, r.reason, content)


class GetFailed(Exception):
    """Raised when a request fails."""

    def __init__(self, status, reason, content):
        """Initialize a new instance of GetFailed."""
        super().__init__(f"{status} {reason}\n{content})")